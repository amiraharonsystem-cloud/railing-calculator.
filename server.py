import os
import io
from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)
CORS(app)

# פונקציית עזר לכתיבה לתא (גם אם הוא ממוזג)
def write_to_cell(ws, cell_address, value):
    cell = ws[cell_address]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_address in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    cell.value = value

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        template_path = 'template.xlsx'
        
        # טעינת התבנית
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
        else:
            # אם אין תבנית, יוצר קובץ חדש כדי שלא יקרוס
            wb = Workbook()
            
        ws = wb.active
        
        # --- שתילת הנתונים בתאים (לפי הפורמט שלך) ---
        # תאריך
        write_to_cell(ws, 'B2', data.get('date', ''))
        # שם המבודק (קבוע)
        write_to_cell(ws, 'E2', "אמיר אהרון")
        # שם הפרויקט / קוד פרויקט
        write_to_cell(ws, 'B3', data.get('project', ''))
        # מספר הזמנה
        write_to_cell(ws, 'E3', data.get('order', ''))
        # כתובת האתר
        write_to_cell(ws, 'B4', data.get('address', ''))

        # הוספת סכומים קבועים (אם רלוונטי לדוח שלך)
        # write_to_cell(ws, 'F15', 1693.68)

        # הכנת הקובץ למשלוח
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name="report.xlsx"
        )
    except Exception as e:
        print(f"Error: {e}")
        return str(e), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)
